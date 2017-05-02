from flask import Flask,url_for,render_template,session,g,redirect,abort,flash
from flask import request,make_response
from werkzeug import secure_filename
from openpyxl import load_workbook
import os
import sqlite3
import pymysql


app = Flask(__name__)
app.config.from_object(__name__)
app.config.update(dict(
    DATABASE=os.path.join(app.root_path, 'flask.db'),
    SECRET_KEY='development key',
    USERNAME='admin',
    PASSWORD='admin'
))
app.config.from_envvar('FLASKR_SETTINGS', silent=True)

##DB
#mysql
conn = pymysql.connect(host='localhost', port=3307, user='root', passwd='123456', db='mydb', charset='utf8')
cur=conn.cursor()
cur.execute('select * from people')
data=cur.fetchall()
for d in data:
    print("id:"+str(d[0])+"name:"+d[1])
cur.close()
#conn.close()

##excel
#read excel
@app.route('/rxlsx/')
def rxlsx():
    #windows
    #wb1=load_workbook('E:/work/flask/flask-web/upload/451.1.IDCV6.2.1.xlsx')
    #mac
    wb1=load_workbook('/Users/xuzhengxi/Desktop/test.xlsx')
    st=wb1['设备表']
    # sheetnames=wb1.get_sheet_names()
    # st=wb1.get_sheet_by_name(sheetnames[1])
    #data=st['A1'].value
    # data_dic={}
    # for rx in range(1,10):
    #     temp_list=[]
    #     pid = st.cell(row = rx,column = 1 ).value
    #     w1 = st.cell(row = rx,column = 2 ).value
    #     w2 = st.cell(row = rx,column = 3 ).value
    #     w3 = st.cell(row = rx,column = 4 ).value
    #     w4 = st.cell(row = rx,column = 5 ).value
    #     temp_list=[pid,w1,w2,w3,w4]
    #     data_dic[pid]=temp_list
    list=[]
    for row in st['A1':'AP']:
        for cell in row:
            list.append(cell.value)

    return render_template('rxlsx.html',data=list)

#insert excel data to mysql
@app.route('/saveexcel/')
def save_excel():
    #conn = pymysql.connect(host='localhost', port=3307, user='root', passwd='123456', db='mydb', charset='utf8')
    curexcel=conn.cursor()
    insert_sql='INSERT INTO idc (id, idc, hostname, company_num, place, telecom_ip) VALUES (%s, %s, %s, %s, %s, %s)'
    wb=load_workbook('/Users/xuzhengxi/PycharmProjects/flask-web/upload/test.xlsx')
    st=wb['设备表']
    excel_data=[]
    for row in st.rows:
        for cell in row:
            excel_data.append(cell.value)
        curexcel.execute(insert_sql,(excel_data[0],excel_data[1],excel_data[2],excel_data[3],excel_data[4],excel_data[5]))
        excel_data=[]
    conn.commit()
    curexcel.close()
    conn.close()



##DB
#connect DB
def connect_db():
    rv=sqlite3.connect(app.config['DATABASE'])
    rv.row_factory=sqlite3.Row
    return rv

#get DB
def get_db():
    if not hasattr(g,'sqlite_db'):
        g.sqlite_db=connect_db()
    return g.sqlite_db

#init DB
def init_db():
    db=get_db()
    with app.open_resource('schema.sql','r') as f:
        db.cursor().executescript(f.read())
    db.commit()

@app.cli.command('initdb')
def initdb_command():
    init_db()
    print('Initialized the database.')

##show entrites
@app.route('/showen/')
def show_entries():
    db=get_db()
    cur=db.execute('select title, text from entries order by id desc')
    entries=cur.fetchall()
    return render_template('show_entries.html',entries=entries)

##add entrites
@app.route('/adden/',methods=['GET','POST'])
def add_entry():
    if not session.get('logged_in'):
        abort(401)
    db=get_db()
    db.execute('insert into entries (title, text) values (?, ?)',[request.form['title'],request.form['text']])
    db.commit()
    flash('New entry was successfully posted')
    return redirect(url_for('show_entries'))

##login
@app.route('/login/',methods=['GET','POST'])
def login():
    error=None
    if request.method=='POST':
        if request.form['username']!=app.config['USERNAME']:
            error='Invalid username'
        elif request.form['password']!=app.config['PASSWORD']:
            error='Invalid password'
        else:
            session['logged_in']=True
            flash('You were logged in')
            return redirect(url_for('show_entries'))
    return render_template('login.html',error=error)

##logout
@app.route('/logout/')
def logout():
    session.pop('logged_in',None)
    flash('You were logged out')
    return redirect(url_for('show_entries'))

##router
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/hello/<name>')
def hello_world(name):
    return render_template('hello.html',name=name)

## R/W cookies
@app.route('/wck/<username>/')
def write_cookies(username):
    resp=make_response(render_template('index.html'))
    resp.set_cookie('username',username)
    return resp

@app.route('/rck/')
def read_cookies():
    username=request.cookies.get('username')
    return username

##构造url
with app.test_request_context():
 url_for('static',filename='style.css')

###file upload
@app.route('/upload/',methods=['GET','POST'])
def upload_file():
    if request.method=='GET':
        return render_template('upload.html')
    if request.method=='POST':
        f=request.files['the_file']
        #f.save('/Users/xuzhengxi/PycharmProjects/flask-web/upload/upload_file.txt')
        filename=secure_filename(f.filename)
        #windows
        #fpath="E:/work/flask/flask-web/upload/"+filename
        #mac
        fpath="/Users/xuzhengxi/PycharmProjects/flask-web/upload/"+filename
        f.save(fpath)
        return 'upload success.'



if __name__ == '__main__':
    app.debug=True
    app.run()
